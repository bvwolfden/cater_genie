#!/usr/bin/env python3
"""Extract the client's operational spreadsheets into clean JSON seed files.

Output -> prisma/seed-data/*.json, consumed by scripts/seed.ts.

This is intentionally tolerant: real-world ops sheets contain #DIV/0!, blank
rows, and typo'd date ranges. We coerce gently and skip junk, printing counts.
"""
import json
import os
import re
import datetime as dt
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "prisma", "seed-data")
os.makedirs(OUT, exist_ok=True)

DAILY_XLSX = os.path.join(ROOT, "Daily Dash Board Online.xlsx")
TIMESHEET_XLSX = os.path.join(ROOT, "Timesheets - Jun 15 - Jun 21, 2026 (2).xlsx")
COMP_XLSX = os.path.join(ROOT, "Weekly Sales Comp 25_26.xlsx")
PROJ_XLSX = os.path.join(ROOT, "weekly projections 2026.xlsx")


def num(v):
    """Coerce a cell to a rounded float, or None for blanks/errors."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        return round(float(v), 4)
    s = str(v).strip()
    if not s or s.startswith("#"):
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "")
    try:
        return round(float(s), 4)
    except ValueError:
        return None


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    return None


def header_map(ws, header_row=1):
    """Map normalized header text -> 1-based column index."""
    m = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if h is not None:
            key = re.sub(r"\s+", " ", str(h).strip().lower())
            m[key] = c
    return m


def find_col(hmap, *needles):
    """Find a column whose header contains all needle substrings."""
    for key, col in hmap.items():
        if all(n in key for n in needles):
            return col
    return None


def parse_week_start(s, default_year=2026):
    """Parse the START date of a week-range string.

    Handles '1/1/26 - 1/4/26', '4/13/26 -4/19/2', '1/1-1/4', '3/2-3-8'.
    Returns (start_iso, end_iso|None).
    """
    if not s:
        return None, None
    s = str(s).strip()
    # split into start / end on the first range separator (space-dash-space, or
    # a dash between two m/d tokens)
    parts = re.split(r"\s*[-–]\s*", s)
    start_raw = parts[0].strip()
    end_raw = parts[1].strip() if len(parts) > 1 else None

    def one(token, year):
        token = token.strip().strip("-")
        bits = re.findall(r"\d+", token)
        if len(bits) < 2:
            return None
        mo, da = int(bits[0]), int(bits[1])
        yr = year
        if len(bits) >= 3:
            yr = int(bits[2])
            if yr < 100:
                yr += 2000
        try:
            return dt.date(yr, mo, da)
        except ValueError:
            return None

    start = one(start_raw, default_year)
    end = one(end_raw, start.year if start else default_year) if end_raw else None
    return (start.isoformat() if start else None,
            end.isoformat() if end else None)


def write(name, rows):
    path = os.path.join(OUT, name)
    with open(path, "w") as f:
        json.dump(rows, f, indent=2, default=str)
    print(f"  wrote {len(rows):>4} rows -> prisma/seed-data/{name}")


# ---------------------------------------------------------------------------
# 1. Daily Tracker -> daily_metrics + account_balances
# ---------------------------------------------------------------------------
def extract_daily():
    wb = load_workbook(DAILY_XLSX, data_only=True)
    ws = wb["Daily Tracker (2)"]
    h = header_map(ws)

    c_week = find_col(h, "week")
    c_date = find_col(h, "date")
    c_net = find_col(h, "net", "sales")
    c_tax = find_col(h, "tax")
    c_labor = find_col(h, "labor", "$")
    c_pct = find_col(h, "labor", "% daily") or find_col(h, "% daily")
    c_hours = find_col(h, "total", "hours")
    c_food = find_col(h, "food")
    c_notes = find_col(h, "notes")

    acct_cols = {
        "OPERATING": find_col(h, "op", "acct"),
        "PAYROLL": find_col(h, "payroll"),
        "MERCHANT": find_col(h, "merchant"),
        "SAVINGS": find_col(h, "savings"),
        "HOLDING": find_col(h, "holding"),
        "CC_PROCESSING": find_col(h, "cc", "processing") or find_col(h, "processing"),
    }

    metrics, balances = [], []
    for r in range(2, ws.max_row + 1):
        date = as_date(ws.cell(row=r, column=c_date).value) if c_date else None
        if not date:
            continue
        net = num(ws.cell(row=r, column=c_net).value) if c_net else None
        tax = num(ws.cell(row=r, column=c_tax).value) if c_tax else None
        labor = num(ws.cell(row=r, column=c_labor).value) if c_labor else None
        pct = num(ws.cell(row=r, column=c_pct).value) if c_pct else None
        hours = num(ws.cell(row=r, column=c_hours).value) if c_hours else None
        food = num(ws.cell(row=r, column=c_food).value) if c_food else None
        notes = text(ws.cell(row=r, column=c_notes).value) if c_notes else None

        acct_vals = {}
        for acct, col in acct_cols.items():
            if col:
                v = num(ws.cell(row=r, column=col).value)
                if v is not None:
                    acct_vals[acct] = v

        has_any = any(x is not None for x in (net, tax, labor, hours, food)) or acct_vals
        if not has_any:
            continue

        metrics.append({
            "date": date,
            "weekLabel": text(ws.cell(row=r, column=c_week).value) if c_week else None,
            "netSales": net,
            "tax": tax,
            "laborCost": labor,
            "laborHours": hours,
            "laborPct": pct,
            "foodPurchases": food,
            "notes": notes,
        })
        for acct, v in acct_vals.items():
            balances.append({"date": date, "account": acct, "balance": v})

    write("daily_metrics.json", metrics)
    write("account_balances.json", balances)


# ---------------------------------------------------------------------------
# 2. Timesheets -> labor_entries
# ---------------------------------------------------------------------------
def extract_labor():
    wb = load_workbook(TIMESHEET_XLSX, data_only=True)
    ws = wb["Entries"]
    h = header_map(ws)

    c = {
        "first": find_col(h, "first"),
        "last": find_col(h, "last"),
        "eid": find_col(h, "employee", "id"),
        "date": find_col(h, "date"),
        "regular": find_col(h, "regular"),
        "rate": find_col(h, "hourly", "rate"),
        "ot": find_col(h, "ot") if find_col(h, "ot") and "double" not in next((k for k in h if h[k] == find_col(h, "ot")), "") else None,
        "dot": find_col(h, "double"),
        "paid": find_col(h, "paid", "total"),
        "tips": find_col(h, "tips"),
        "earn": find_col(h, "earnings"),
        "sched": find_col(h, "schedule"),
        "site": find_col(h, "job", "site"),
        "pos": find_col(h, "position"),
    }
    # plain "OT" column (not "Double OT")
    c_ot = None
    for key, col in h.items():
        if key == "ot":
            c_ot = col

    rows = []
    for r in range(2, ws.max_row + 1):
        date = as_date(ws.cell(row=r, column=c["date"]).value) if c["date"] else None
        if not date:
            continue
        rows.append({
            "date": date,
            "firstName": text(ws.cell(row=r, column=c["first"]).value) if c["first"] else None,
            "lastName": text(ws.cell(row=r, column=c["last"]).value) if c["last"] else None,
            "employeeId": text(ws.cell(row=r, column=c["eid"]).value) if c["eid"] else None,
            "department": text(ws.cell(row=r, column=c["sched"]).value) if c["sched"] else None,
            "position": text(ws.cell(row=r, column=c["pos"]).value) if c["pos"] else None,
            "jobSite": text(ws.cell(row=r, column=c["site"]).value) if c["site"] else None,
            "regularHours": num(ws.cell(row=r, column=c["regular"]).value) if c["regular"] else None,
            "otHours": num(ws.cell(row=r, column=c_ot).value) if c_ot else None,
            "doubleOtHours": num(ws.cell(row=r, column=c["dot"]).value) if c["dot"] else None,
            "hourlyRate": num(ws.cell(row=r, column=c["rate"]).value) if c["rate"] else None,
            "paidTotal": num(ws.cell(row=r, column=c["paid"]).value) if c["paid"] else None,
            "tips": num(ws.cell(row=r, column=c["tips"]).value) if c["tips"] else None,
            "earningsTotal": num(ws.cell(row=r, column=c["earn"]).value) if c["earn"] else None,
        })
    write("labor_entries.json", rows)


# ---------------------------------------------------------------------------
# 3. Weekly Sales Comp + Projections -> weekly_rollup + weekly_channel_revenue
# ---------------------------------------------------------------------------
def extract_weekly():
    wb = load_workbook(COMP_XLSX, data_only=True)
    ws = wb["Revenue Comp 25_24_23"]
    h = header_map(ws)

    c_week = find_col(h, "week")
    c_2023 = find_col(h, "2023")
    c_2025 = find_col(h, "revenue 2025")
    c_2026 = find_col(h, "revenue 2026")
    c_labor = find_col(h, "gross", "weekly")  # (Gross) Weekly Labor
    c_pct = find_col(h, "labor %", "2026")
    c_ease = find_col(h, "ease", "subtotal")
    c_trax = find_col(h, "trax", "subtotal")
    c_cafe = find_col(h, "cafe", "net")

    rollups = {}
    channel = {}  # (weekStart, channel) -> {actual, projected, weekEnd}

    def chan(ws_, ch, key, weekStart, weekEnd, val, kind):
        if val is None:
            return
        k = (weekStart, ch)
        d = channel.setdefault(k, {"weekStart": weekStart, "weekEnd": weekEnd, "channel": ch, "actual": None, "projected": None})
        d[kind] = val
        if weekEnd and not d.get("weekEnd"):
            d["weekEnd"] = weekEnd

    for r in range(2, ws.max_row + 1):
        wk = ws.cell(row=r, column=c_week).value if c_week else None
        start, end = parse_week_start(wk)
        if not start:
            continue
        total = num(ws.cell(row=r, column=c_2026).value) if c_2026 else None
        prev1 = num(ws.cell(row=r, column=c_2025).value) if c_2025 else None
        prev3 = num(ws.cell(row=r, column=c_2023).value) if c_2023 else None
        labor = num(ws.cell(row=r, column=c_labor).value) if c_labor else None
        pct = num(ws.cell(row=r, column=c_pct).value) if c_pct else None
        if any(x is not None for x in (total, prev1, prev3, labor, pct)):
            rollups[start] = {
                "weekStart": start, "weekEnd": end,
                "totalRevenue": total, "revenuePrev1": prev1,
                "revenuePrev2": None, "revenuePrev3": prev3,
                "laborCost": labor, "laborPct": pct, "projectedTotal": None,
            }
        chan(ws, "CATEREASE", c_ease, start, end, num(ws.cell(row=r, column=c_ease).value) if c_ease else None, "actual")
        chan(ws, "CATERTRAX", c_trax, start, end, num(ws.cell(row=r, column=c_trax).value) if c_trax else None, "actual")
        chan(ws, "CAFE_RETAIL", c_cafe, start, end, num(ws.cell(row=r, column=c_cafe).value) if c_cafe else None, "actual")

    # projections workbook
    pwb = load_workbook(PROJ_XLSX, data_only=True)
    pws = pwb["Sheet1"]
    # header rows are messy; channel labels live around row 3: caterease/catertrax/aloha
    # data rows: col1 = range, col2 = caterease, col4 = catertrax, col6 = aloha, col8 = total
    for r in range(4, pws.max_row + 1):
        rng = pws.cell(row=r, column=1).value
        start, end = parse_week_start(rng)
        if not start:
            continue
        ease = num(pws.cell(row=r, column=2).value)
        trax = num(pws.cell(row=r, column=4).value)
        aloha = num(pws.cell(row=r, column=6).value)
        total = num(pws.cell(row=r, column=8).value)
        chan(pws, "CATEREASE", 2, start, end, ease, "projected")
        chan(pws, "CATERTRAX", 4, start, end, trax, "projected")
        chan(pws, "ALOHA", 6, start, end, aloha, "projected")
        if total is not None:
            if start in rollups:
                rollups[start]["projectedTotal"] = total
            else:
                rollups[start] = {
                    "weekStart": start, "weekEnd": end, "totalRevenue": None,
                    "revenuePrev1": None, "revenuePrev2": None, "revenuePrev3": None,
                    "laborCost": None, "laborPct": None, "projectedTotal": total,
                }

    write("weekly_rollup.json", list(rollups.values()))
    write("weekly_channel_revenue.json", list(channel.values()))


if __name__ == "__main__":
    print("Extracting spreadsheets -> prisma/seed-data/")
    extract_daily()
    extract_labor()
    extract_weekly()
    print("Done.")
